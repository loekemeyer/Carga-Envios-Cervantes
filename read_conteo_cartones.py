# -*- coding: utf-8 -*-
"""Lee el conteo cartones del Excel y exporta JSON con cod -> stock"""
import json
from openpyxl import load_workbook

PATH = "S:/AA  LOGISTICA/A4 Relevamiento INSUMOS Stock y OP/A1 Conteo INSUMOS y OC (Compra)/A3 Cartones/A2 Conteo Cartones VIGENTE/Conteo Pedido Cartones 07-04-26.xlsx"
SHEET = "Conteo 07-26"

wb = load_workbook(PATH, data_only=True)
print("Sheets:", wb.sheetnames)
if SHEET not in wb.sheetnames:
    print(f"Sheet '{SHEET}' no encontrada")
    raise SystemExit(1)

ws = wb[SHEET]

# Header en row 7 → datos desde row 8.
# Cols: A=Cod (idx 0), B=L (linea), C=Descripcion, D=Sector, J=Stock Uni/Paq/KG (idx 9)
out = {}
for row in ws.iter_rows(min_row=8, values_only=True):
    cod = row[0]
    linea = row[1]
    desc = row[2]
    stock = row[9]
    if cod is None: continue
    cod_str = str(cod).strip()
    if not cod_str or cod_str.startswith("#"): continue
    # Filtrar codigos no validos (header repeats, descripciones)
    if cod_str.lower() in ("cod", "varios"): continue
    if cod_str.startswith("Corb"): continue
    if cod_str.startswith("Bomb"): continue
    try:
        stock_n = int(stock) if stock is not None else 0
    except (ValueError, TypeError):
        stock_n = 0
    out[cod_str] = stock_n

print(f"Total codes: {len(out)}")
print(json.dumps(out, indent=2, ensure_ascii=False)[:3000])

# Generar SQL
sql_values = ",".join([f"('{c}',{s})" for c, s in out.items()])
print("\n\n--- SQL ---")
print(f"INSERT INTO \"Stock_Inicial_Cartones\" (cod, stock_inicial) VALUES\n{sql_values}\nON CONFLICT (cod) DO UPDATE SET stock_inicial=EXCLUDED.stock_inicial, fecha_conteo='2026-05-07';")

