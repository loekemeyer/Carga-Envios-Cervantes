# -*- coding: utf-8 -*-
"""Lee conteo Bombilla 5-05 del .xlsx"""
from openpyxl import load_workbook

PATH = "S:/AA  LOGISTICA/A4 Relevamiento INSUMOS Stock y OP/A1 Conteo INSUMOS y OC (Compra)/A6 Bombillas/A2 Conteo Bombillas VIGENTE/Conteo Bombilla 5-05-26.xlsx"

wb = load_workbook(PATH, data_only=True)
print("Sheets:", wb.sheetnames)

ws = wb["Conteo 5-05"]
print(f"Filas: {ws.max_row}")
# Mostrar todas las filas con descripcion no vacia
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    desc = str(row[2] or "").strip()
    if desc and not desc.startswith("Este pedido") and "Fecha" not in desc and "Descripcion" not in desc and "C" != desc[:1] or "Limpia" in desc or "Cepillo" in desc:
        print(i, row[0], "|", desc, "|", "Stock=", row[10])
