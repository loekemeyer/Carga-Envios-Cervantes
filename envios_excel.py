# -*- coding: utf-8 -*-
"""Genera Excel de envios por tallerista, una fila por envio.
Incluye comparacion vs promedio del articulo para detectar errores de tipeo."""
import urllib.request
import urllib.parse
import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

URL = "https://hrxfctzncixxqmpfhskv.supabase.co"
KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImhyeGZjdHpuY2l4eHFtcGZoc2t2Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MjQyNjEsImV4cCI6MjA4ODMwMDI2MX0.4L6wguch8UZGhC2VpzrWcCjJGUV-IkYsl9JoCWrOLUs"

def fetch_all():
    rows = []
    page = 0
    page_size = 1000
    while True:
        params = {
            "select": "id,Dia-mes,Tallerista,Sector,Descripcion,KG,Cajones,Unidades",
            "KG": "not.is.null",
            "Cajones": "not.is.null",
            "order": "Tallerista.asc,Descripcion.asc,Dia-mes.asc",
        }
        qs = urllib.parse.urlencode(params)
        req = urllib.request.Request(
            f"{URL}/rest/v1/Envios%20a%20Talleristas?{qs}",
            headers={
                "apikey": KEY,
                "Authorization": f"Bearer {KEY}",
                "Range-Unit": "items",
                "Range": f"{page*page_size}-{(page+1)*page_size-1}",
            }
        )
        with urllib.request.urlopen(req) as resp:
            chunk = json.loads(resp.read().decode("utf-8"))
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        page += 1
    return rows

def main():
    print("Bajando envios...")
    rows = fetch_all()
    print(f"Total: {len(rows)} filas")

    # Solo validos
    rows = [r for r in rows if r.get("Cajones") and float(r["Cajones"]) > 0 and r.get("KG") and float(r["KG"]) > 0]
    print(f"Validos (kg>0 y cajones>0): {len(rows)}")

    # Calcular promedio kg/cajon por (tallerista, descripcion) usando totales (mas estable que media de ratios)
    tot_kg = defaultdict(float)
    tot_caj = defaultdict(float)
    cuenta = defaultdict(int)
    for r in rows:
        key = (r.get("Tallerista") or "", r.get("Descripcion") or "")
        tot_kg[key] += float(r["KG"])
        tot_caj[key] += float(r["Cajones"])
        cuenta[key] += 1

    promedios = {k: (tot_kg[k] / tot_caj[k]) if tot_caj[k] else 0 for k in tot_kg}

    # Limpiar saltos de linea en descripciones para sort y display
    def clean_desc(d):
        return (d or "").replace("\n", " ").replace("\r", " ").strip()

    # Ordenar: tallerista, descripcion limpia, dia-mes
    rows.sort(key=lambda r: (r.get("Tallerista") or "", clean_desc(r.get("Descripcion")), r.get("Dia-mes") or ""))

    # Talleristas con envios
    talleristas = sorted(set(r["Tallerista"] for r in rows if r.get("Tallerista")))
    print(f"Talleristas con envios: {len(talleristas)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Envios"

    headers = ["id", "Dia-mes", "Tallerista", "Sector", "Descripcion", "KG", "Cajones", "Unidades",
               "KG/Cajon", "Prom Art.", "N envios art.", "Desvio % vs prom"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    rojo = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    amarillo = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    fila_actual = 1
    for r in rows:
        kg = float(r["KG"])
        caj = float(r["Cajones"])
        ratio = kg / caj if caj else 0
        key = (r.get("Tallerista") or "", r.get("Descripcion") or "")
        prom = promedios.get(key, 0)
        n_env = cuenta.get(key, 0)
        desv_pct = ((ratio - prom) / prom * 100) if prom else 0

        ws.append([
            r.get("id"),
            r.get("Dia-mes"),
            r.get("Tallerista"),
            r.get("Sector"),
            clean_desc(r.get("Descripcion")),
            kg,
            caj,
            r.get("Unidades"),
            round(ratio, 2),
            round(prom, 2),
            n_env,
            round(desv_pct, 1),
        ])
        fila_actual += 1

        # Pintar fila si desvio absoluto > 25% (rojo) o > 15% (amarillo). Solo si hay >=2 envios.
        if n_env >= 2:
            abs_desv = abs(desv_pct)
            if abs_desv > 25:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=fila_actual, column=col_idx).fill = rojo
            elif abs_desv > 15:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=fila_actual, column=col_idx).fill = amarillo

    # Anchos de columna
    anchos = [("A", 8), ("B", 12), ("C", 14), ("D", 10), ("E", 40), ("F", 10), ("G", 10), ("H", 10),
              ("I", 11), ("J", 11), ("K", 11), ("L", 16)]
    for col, ancho in anchos:
        ws.column_dimensions[col].width = ancho

    # Freeze first row + first 5 cols (id, fecha, tallerista, sector, descripcion)
    ws.freeze_panes = "F2"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Envios_Talleristas_KG_Cajones.xlsx")
    wb.save(out)
    print(f"Guardado: {out}")
    print(f"Talleristas: {', '.join(talleristas)}")
    print("Filas con desvio >25% pintadas en naranja, 15-25% en amarillo (solo si articulo tiene >=2 envios).")

if __name__ == "__main__":
    main()
