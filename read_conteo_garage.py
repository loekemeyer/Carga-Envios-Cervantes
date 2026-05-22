# -*- coding: utf-8 -*-
"""Lee conteo Garage 05-05 del .xlsx"""
from openpyxl import load_workbook

PATH = "S:/AA  LOGISTICA/A4 Relevamiento INSUMOS Stock y OP/A1 Conteo INSUMOS y OC (Compra)/A7 Garage/Relevamiento Garage Vigente.xlsx"
SHEET = "Garage Vacio 2026"

wb = load_workbook(PATH, data_only=True)
print("Sheets:", wb.sheetnames)
if SHEET not in wb.sheetnames:
    print(f"Sheet '{SHEET}' no encontrada")
    raise SystemExit(1)

ws = wb[SHEET]
print(f"Filas: {ws.max_row}, Cols: {ws.max_column}")
print("\n--- Primeras 30 filas ---")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 30: break
    print(i, row[:15])
